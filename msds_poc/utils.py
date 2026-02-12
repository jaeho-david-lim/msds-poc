"""Utility functions for MSDS PoC."""

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pdfplumber
import pandas as pd

logger = logging.getLogger(__name__)


def load_config(config_path: Path) -> Dict[str, Any]:
    """Load configuration from JSON file."""
    with open(config_path, "r") as f:
        return json.load(f)


def save_result(output_path: Path, data: Dict[str, Any]) -> None:
    """Save result to JSON file with UTF-8 encoding and proper formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


def extract_text_from_pdf(pdf_path: Path) -> str:
    """
    Extract all text from PDF file with proper encoding handling.
    
    Args:
        pdf_path: Path to PDF file
        
    Returns:
        Extracted text as string
    """
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            logger.info(f"Processing PDF: {pdf_path.name} ({len(pdf.pages)} pages)")
            for page_num, page in enumerate(pdf.pages, 1):
                try:
                    page_text = page.extract_text(layout=False)
                    if page_text:
                        text += f"\n--- Page {page_num} ---\n{page_text}"
                except Exception as e:
                    logger.warning(f"Error extracting text from page {page_num}: {e}")
                    continue
        
        logger.info(f"Extracted {len(text)} characters from {pdf_path.name}")
    except Exception as e:
        logger.error(f"Error extracting text from {pdf_path}: {e}")
        
    return text


def parse_msds_data(text: str) -> Dict[str, str]:
    """
    Parse MSDS information from extracted text.
    Handles both English and Korean text.
    
    Args:
        text: Extracted text from PDF
        
    Returns:
        Dictionary with parsed MSDS data
    """
    # Common MSDS section headers (English and Korean)
    msds_sections = {
        "Product Name / 제품명": [],
        "Supplier / 공급자": [],
        "Chemical Name / 화학명": [],
        "CAS Number / CAS 번호": [],
        "Hazards / 위험": [],
        "Physical Properties / 물리적 성질": [],
        "Chemical Properties / 화학적 성질": [],
        "Safety Information / 안전 정보": [],
        "Storage / 보관": [],
        "Disposal / 폐기": [],
        "PPE / 개인보호장비": [],
        "First Aid / 응급처치": [],
    }
    
    # Simple parsing logic - extract lines containing common keywords
    lines = text.split('\n')
    current_section = None
    
    for line in lines:
        line_lower = line.lower().strip()
        
        # Check for section headers (English and Korean)
        if any(x in line_lower for x in ['product name', 'product identifier', '제품명', '제품 이름']):
            current_section = "Product Name / 제품명"
        elif any(x in line_lower for x in ['supplier', 'manufacturer', '공급자', '제조사']):
            current_section = "Supplier / 공급자"
        elif any(x in line_lower for x in ['chemical name', '화학명', 'chemical  name']):
            current_section = "Chemical Name / 화학명"
        elif any(x in line_lower for x in ['cas number', 'cas no', 'cas', 'cas 번호']):
            current_section = "CAS Number / CAS 번호"
        elif any(x in line_lower for x in ['hazard', 'danger', '위험', '위험성']):
            current_section = "Hazards / 위험"
        elif any(x in line_lower for x in ['physical property', 'appearance', '물리적 성질', '외관']):
            current_section = "Physical Properties / 물리적 성질"
        elif any(x in line_lower for x in ['chemical property', '화학적 성질']):
            current_section = "Chemical Properties / 화학적 성질"
        elif any(x in line_lower for x in ['safety', '안전']):
            current_section = "Safety Information / 안전 정보"
        elif any(x in line_lower for x in ['storage', '보관']):
            current_section = "Storage / 보관"
        elif any(x in line_lower for x in ['disposal', '폐기']):
            current_section = "Disposal / 폐기"
        elif any(x in line_lower for x in ['ppe', 'protective equipment', 'precaution', '개인보호장비', '보호']):
            current_section = "PPE / 개인보호장비"
        elif any(x in line_lower for x in ['first aid', '응급처치']):
            current_section = "First Aid / 응급처치"
        elif current_section and line.strip() and len(line.strip()) > 2:
            # Add non-empty lines to current section
            msds_sections[current_section].append(line.strip())
    
    # Convert lists to comma-separated strings
    parsed_data = {}
    for key, values in msds_sections.items():
        parsed_data[key] = " | ".join(values[:5]) if values else "N/A"  # Limit to first 5 lines
    
    return parsed_data


def create_excel_from_msds(msds_data: Dict[str, str], output_path: Path) -> None:
    """
    Create Excel workbook from parsed MSDS data.
    
    Args:
        msds_data: Dictionary with parsed MSDS data
        output_path: Path to save Excel file
    """
    try:
        from openpyxl.styles import Font
        
        # Create DataFrame
        df = pd.DataFrame(list(msds_data.items()), columns=['Field', 'Value'])
        
        # Create workbook with formatting
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='MSDS Data', index=False)
            
            # Format worksheet
            worksheet = writer.sheets['MSDS Data']
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 100
            
            # Set font for Korean support
            font = Font(name='Calibri', size=11)
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                          min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.font = font
            
        logger.info(f"Excel file created: {output_path}")
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")


def process_pdf_files(input_dir: Path, output_dir: Path) -> List[Dict[str, Any]]:
    """
    Process all PDF files in input directory and create Excel outputs.
    Supports both English and Korean text.
    
    Args:
        input_dir: Directory containing PDF files
        output_dir: Directory to save Excel files
        
    Returns:
        List of processed file information
    """
    results = []
    
    pdf_files = list(input_dir.glob("*.pdf"))
    if not pdf_files:
        logger.warning(f"No PDF files found in {input_dir}")
        return results
    
    logger.info(f"Found {len(pdf_files)} PDF file(s) to process")
    
    for pdf_file in pdf_files:
        logger.info(f"Processing: {pdf_file.name}")
        
        try:
            # Extract text
            extracted_text = extract_text_from_pdf(pdf_file)
            
            if not extracted_text:
                logger.warning(f"No text extracted from {pdf_file.name}")
                continue
            
            # Parse MSDS data
            msds_data = parse_msds_data(extracted_text)
            
            # Save raw extracted text as backup with correct encoding
            text_output = output_dir / f"{pdf_file.stem}_extracted.txt"
            text_output.parent.mkdir(parents=True, exist_ok=True)
            with open(text_output, 'w', encoding='utf-8') as f:
                f.write(extracted_text)
            
            # Create Excel file
            excel_output = output_dir / f"{pdf_file.stem}.xlsx"
            create_excel_from_msds(msds_data, excel_output)
            
            results.append({
                "input_file": pdf_file.name,
                "output_excel": excel_output.name,
                "output_text": text_output.name,
                "status": "success",
                "message": "PDF 처리 완료 / PDF processed successfully"
            })
            
            logger.info(f"✓ Successfully processed: {pdf_file.name}")
            
        except Exception as e:
            logger.error(f"Failed to process {pdf_file.name}: {e}")
            results.append({
                "input_file": pdf_file.name,
                "status": "failed",
                "error": str(e),
                "message": f"처리 실패 / Processing failed: {e}"
            })
    
    return results

